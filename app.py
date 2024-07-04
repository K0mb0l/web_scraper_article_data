import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from email import encoders
import smtplib
import aiohttp
import asyncio
import os
import re

# Anmeldedaten und URLs
username = 'info@caipi.de'
password = 'tca.vxn7ztr8ypr6PAX'
login_url = 'https://www.nebelung.de/account/login'
shop_url = 'https://www.nebelung.de/shop/'
base_url = 'https://www.nebelung.de'

# Email-Konfiguration
email_user = 'webscrapertestnebelungen@gmail.com'
email_password = 'orxa gayk uurh cdkk'
email_send_to = 'steglichmaximilian@gmail.com'
email_subject = 'Nebelung Artikel Update'
smtp_server = 'smtp.gmail.com'
smtp_port = 465  # Using SSL port

# Funktion zum Abrufen der Anmeldeseite und Extrahieren versteckter Eingabefelder
async def get_hidden_inputs(session):
    async with session.get(login_url) as response:
        soup = BeautifulSoup(await response.text(), 'html.parser')
        hidden_inputs = soup.find_all('input', type='hidden')
        login_payload = {hidden_input.get('name'): hidden_input.get('value') for hidden_input in hidden_inputs if hidden_input.get('name') and hidden_input.get('value')}
        login_payload.update({'email': username, 'password': password})
        return login_payload

# Funktion zum Anmelden und Abrufen von Kategorie-URLs
async def login_and_get_category_urls(session, login_payload):
    async with session.post(login_url, data=login_payload) as response:
        if response.status != 200 or "account/logout" not in await response.text():
            print("Fehler bei der Anmeldung.")
            exit()
    
    async with session.get(shop_url) as response:
        soup = BeautifulSoup(await response.text(), 'html.parser')
        category_links = soup.find_all('a', class_='btn btn-primary btn-md btn-block')
        return [base_url + link.get('href') for link in category_links][:-2]

async def get_product_urls(session, category_urls):
    product_urls = []
    for url in category_urls:
        print(url)
        page_number = 1  # Seite initialisieren
        while True:
            async with session.get(f"{url}?order=standardsortierung&p={page_number}") as response:
                if response.status != 200:
                    break

                soup = BeautifulSoup(await response.text(), 'html.parser')
                product_links = soup.find_all('a', class_='product-name')
                product_urls.extend([link.get('href') for link in product_links])

                break
                # Überprüfen, ob es eine nächste Seite gibt
                if not soup.find('li', class_='page-item page-next'):
                    break
                page_number += 1  # Seite inkrementieren
    
    return product_urls

# Funktion zum Scrapen von Artikeln aus einer Kategorie-Seite
async def scrape_product_page(session, url):
    print(url)
    products = []
    async with session.get(url) as response:
        if response.status != 200:
            return []

        soup = BeautifulSoup(await response.text(), 'html.parser')
        product_details = soup.find('div', class_='col-lg-5 product-detail-buy')
        product_description = soup.find('div', class_='col-md-7 product-detail-description-text')

        # Get all fields:
        product_manufacturer_a = product_details.find('a', class_='product-detail-manufacturer-link')
        product_name_h1 = product_details.find('h1', class_='product-detail-name')
        product_price_p = product_details.find('p', class_='product-detail-price')
        product_dimension_div = product_details.find('div', class_='icon-information d-flex flex-row')
        product_dimension_span = product_dimension_div.find('span', class_=False) if product_dimension_div else None
        product_number_span = product_details.find('span', class_='product-detail-ordernumber')
        product_number_short_span = product_details.find('div', class_='product-detail-ordernumber-short-container')
        product_content_span = product_details.find('span', class_='price-unit-content')
        product_details_span = product_details.find('span', class_='product-detail-recommended-amount-content')
        product_availabilty_div = product_details.find('div', class_='product-availability-immediatly')
        product_availabilty_span = product_availabilty_div.find('span', class_=False) if product_availabilty_div else None
        product_information_p = product_description.find('p', class_=False)

        product_info = {
            'Product_Manufacturer': product_manufacturer_a.text.strip() if product_manufacturer_a else None,
            'Product_Name': product_name_h1.text.strip() if product_name_h1 else None,
            'Product_Price': product_price_p.text.strip() if product_price_p else None,
            'Product_Dimension': re.sub(r'\s*([x()])\s*', r' \1 ', product_dimension_span.text.strip()) if product_dimension_span else None,
            'Product_Number': product_number_span.text.strip() if product_number_span else None,
            'Product_Number_Short': product_number_short_span.text.replace('Produktnummer (Kurzform):', '').strip() if product_number_short_span else None,
            'Product_Content': product_content_span.text.strip() if product_content_span else None,
            'Product_Recommended_Order_Quantity': product_details_span.text.strip() if product_details_span else None,
            'Product_Availabilty': product_availabilty_span.text.strip() if product_availabilty_span else None,
            'Product_Information': product_information_p.text.strip() if product_information_p else None,
        }
        products.append(product_info)

    return products

# Funktion zum Senden der E-Mail
def send_email(subject, body, file_path=None):
    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = email_send_to
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    if file_path:
        with open(file_path, "rb") as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f"attachment; filename={os.path.basename(file_path)}")
            msg.attach(part)

    try:
        server = smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=30)
        server.login(email_user, email_password)
        text = msg.as_string()
        server.sendmail(email_user, email_send_to, text)
        server.quit()
    except Exception as e:
        print(f"Failed to send email: {e}")

async def main():
    async with aiohttp.ClientSession() as session:
        # Anmeldeinformationen abrufen und anmelden
        login_payload = await get_hidden_inputs(session)
        category_urls = await login_and_get_category_urls(session, login_payload)
        article_urls = await get_product_urls(session, category_urls)

        # Kategorie-Seiten asynchron scrapen
        tasks = [scrape_product_page(session, url) for url in article_urls]
        results = await asyncio.gather(*tasks)
        
        # Produkte konsolidieren
        products = [item for sublist in results for item in sublist]

        # Daten in eine Excel-Datei schreiben oder aktualisieren
        file_path = 'nebelung-artikel-liste-ws.xlsx'
        new_df = pd.DataFrame(products)

        if os.path.exists(file_path):
            # Lade die bestehende Arbeitsmappe
            book = load_workbook(file_path)
            
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                writer._book = book  # Setze das Workbook direkt

                for sheetname in book.sheetnames:
                    existing_df = pd.read_excel(file_path, sheet_name=sheetname)
                    
                    # Debug prints
                    print("Existing DataFrame columns:", existing_df.columns)
                    print("New DataFrame columns:", new_df.columns)
                    
                    # Identifizieren neuer, entfernter und geänderter Artikel
                    new_articles = new_df[~new_df['Product_Number'].isin(existing_df['Product_Number'])]
                    removed_articles = existing_df[~existing_df['Product_Number'].isin(new_df['Product_Number'])]
                    changed_prices = new_df[new_df['Product_Number'].isin(existing_df['Product_Number'])]

                    # Prüfen auf Preisänderungen
                    price_changes = pd.merge(existing_df[['Product_Number', 'Product_Price']], changed_prices[['Product_Number', 'Product_Price']], on='Product_Number')
                    price_changes = price_changes[price_changes['Product_Price_x'] != price_changes['Product_Price_y']]
                    price_changes = pd.merge(price_changes, new_df, on='Product_Number')

                    # Speichern der Excel-Dateien für jede Änderung
                    if not new_articles.empty:
                        new_articles.to_excel('neue_artikel.xlsx', index=False)
                    if not removed_articles.empty:
                        removed_articles.to_excel('entfernte_artikel.xlsx', index=False)
                    if not price_changes.empty:
                        price_changes.to_excel('geanderte_preise.xlsx', index=False)

                    # Aktualisieren der bestehenden Excel-Datei
                    updated_df = pd.concat([existing_df[~existing_df['Product_Number'].isin(removed_articles['Product_Number'])], new_df])
                    updated_df.sort_values(by='Product_Number', inplace=True)
                    updated_df.to_excel(writer, sheet_name=sheetname, index=False)

                # Ergebnisse ausgeben
                body = (
                    f"Neue Artikel: {len(new_articles)}\n"
                    f"Entfernte Artikel: {len(removed_articles)}\n"
                    f"Geänderte Preise: {len(price_changes)}\n"
                )
                send_email(email_subject, body, file_path)

        else:
            new_df.to_excel(file_path, index=False)
            send_email(email_subject, 'Die Excel Datei wurde neu erstellt und angefügt.', file_path)

# Asynchrones Hauptprogramm starten
asyncio.run(main())